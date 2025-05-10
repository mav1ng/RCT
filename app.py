import streamlit as st
import pandas as pd
from grouping_logic.processor import Processor as GroupProcessor
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Randomized Coffee Trial Group Generator", layout="centered", page_icon=":busts_in_silhouette:")

# --- Minimal, clean, modern dark theme ---
st.markdown(
    """
    <style>
    html, body, .main {
        background: #181a1b !important;
        color: #f7f7f7 !important;
    }
    .stApp {
        background: #181a1b !important;
    }
    .stMarkdown, .stTextInput > div > div > input, .stDataFrame, .stTable, .stSelectbox > div > div, .stSelectbox span, .css-1wa3eu0-placeholder, .css-1n76uvr, .css-1d391kg, .css-1cpxqw2, select, option {
        background: #181a1b !important;
        color: #f7f7f7 !important;
        border: none !important;
        box-shadow: none !important;
    }
    .stButton>button {
        color: #fff !important;
        background: #23272a !important;
        border-radius: 6px;
        font-size: 1.1em;
        font-weight: 600;
        border: 1px solid #444 !important;
        margin: 0.5em 0;
        box-shadow: none !important;
    }
    .stDataFrame, .stTable {
        border-radius: 6px;
        font-size: 1.07em;
        background: #23272a !important;
        color: #f7f7f7 !important;
    }
    .stSelectbox [data-baseweb="select"] {
        color: #f7f7f7 !important;
        background: #23272a !important;
    }
    .st-cq {
        background: #23272a !important;
        color: #f7f7f7 !important;
    }
    h1, h2, h3, h4, h5, h6 {
        color: #f7f7f7 !important;
        font-weight: 700;
        letter-spacing: 0.5px;
        border: none;
        margin-bottom: 0.3em;
    }
    .st-bb, .st-b8, .st-bf {
        background: #23272a !important;
        color: #f7f7f7 !important;
        border-radius: 8px;
        box-shadow: none !important;
    }
    .stAlert {
        background: #23272a !important;
        color: #f7f7f7 !important;
        border-radius: 6px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown("<h1 style='text-align:center; color:#f7f7f7; margin-bottom:0.5em;'>Randomized Coffee Trial Group Generator</h1>", unsafe_allow_html=True)

st.markdown("<p style='text-align:center; color:#f7f7f7; font-size:1.15em;'>Upload your participant data and generate diverse, fair groups with a single click!</p>", unsafe_allow_html=True)

def detect_column(columns, keywords):
    """Find the most likely column from list using keyword matching"""
    columns_lower = [col.lower() for col in columns]
    for keyword in keywords:
        for i, col in enumerate(columns_lower):
            if keyword in col:
                return i
    return 0  # Fallback to first column if no matches


def auto_fill_categories(df, position_col, categories, group_size):
    """Auto-fill categories with guaranteed minimum fulfillment"""
    total_participants = len(df)
    min_members = total_participants // group_size  # Ceiling division

    # Save manual assignments
    manual_assignments = {
        cat: df[df[position_col].isin(positions)]
        for cat, positions in categories.items()
    }

    lists = {cat: df_assigned.copy() for cat, df_assigned in manual_assignments.items()}

    def move_participant(source, target):
        """Move one participant between categories"""
        if len(lists[source]) == 0:
            return False
        participant = lists[source].sample(1)
        lists[source] = lists[source].drop(participant.index)
        lists[target] = pd.concat([lists[target], participant])
        return True

    # Phase 1: Priority filling for hierarchical structure
    priority_order = [
        ('Mid-Level', 'Senior-Level'),
        ('Early Career', 'Mid-Level'),
        ('Student', 'Early Career')
    ]

    for source, target in priority_order:
        while len(lists[target]) < min_members:
            if not move_participant(source, target):
                # Fallback to any available source if preferred is empty
                potential_sources = [s for s in categories if len(lists[s]) > 0]
                if not potential_sources:
                    break
                source = max(potential_sources, key=lambda x: len(lists[x]))
                if not move_participant(source, target):
                    break

    # Phase 2: Force minimum requirements
    for cat in categories:
        while len(lists[cat]) < min_members:
            potential_sources = [s for s in categories if len(lists[s]) > min_members]
            if not potential_sources:
                potential_sources = [s for s in categories if len(lists[s]) > 0]
                if not potential_sources:
                    break
            source = max(potential_sources, key=lambda x: len(lists[x]))
            if not move_participant(source, cat):
                break

    # Final validation check
    for cat in categories:
        if len(lists[cat]) < min_members:
            st.error(f"Critical error: {cat} has {len(lists[cat])} members (min {min_members})")
            st.stop()

    return pd.concat(lists.values()), categories


def main():
    with st.container():
        uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

        if uploaded_file:
            df = pd.read_excel(uploaded_file)
            cols = df.columns.tolist()

            # Auto-detect columns with smart defaults
            name_idx = detect_column(cols, ['name', 'full name', 'employee'])
            email_idx = detect_column(cols, ['email', 'mail'])
            position_idx = detect_column(cols, ['position', 'role', 'title'])
            job_sector_idx = detect_column(cols, ['job sector', 'sector', 'department'])

            # Column selection with auto-detected defaults
            col1, col2, col3 = st.columns(3)
            with col1:
                name_col = st.selectbox("Name Column", cols, index=name_idx)
            with col2:
                email_col = st.selectbox("Email Column", cols, index=email_idx)
            with col3:
                job_sector_col = st.selectbox("Job Sector Column", cols, index=job_sector_idx)

            position_col = st.selectbox("Position Column", cols, index=position_idx)

            # Initialize categorization state
            if 'categories' not in st.session_state:
                st.session_state.categories = {
                    'Student': [],
                    'Early Career': [],
                    'Mid-Level': [],
                    'Senior-Level': []
                }

            st.header("Categorize Positions")
            positions = df[position_col].unique().tolist()

            # Calculate unassigned positions
            assigned = sum(st.session_state.categories.values(), [])
            unassigned_positions = [p for p in positions if p not in assigned]

            category_cols = st.columns(4)
            for i, category in enumerate(st.session_state.categories):
                with category_cols[i]:
                    st.markdown(f"<span class='group-label'>{category}</span>", unsafe_allow_html=True)
                    # Only show available positions (not assigned to any group)
                    available = [p for p in positions if p not in assigned or p in st.session_state.categories[category]]
                    # Only allow adding if there are unassigned positions
                    if unassigned_positions:
                        add_pos = st.selectbox(
                            f"Add to {category}",
                            [p for p in unassigned_positions if p not in st.session_state.categories[category]],
                            key=f"add_select_{category}"
                        ) if [p for p in unassigned_positions if p not in st.session_state.categories[category]] else None
                        if add_pos and st.button(f"Add to {category}", key=f"add_btn_{category}"):
                            st.session_state.categories[category].append(add_pos)
                            assigned.append(add_pos)
                            st.rerun()
                    else:
                        st.info("All positions assigned.")
                    # Remove position from category
                    if st.session_state.categories[category]:
                        remove_pos = st.selectbox(
                            f"Remove from {category}",
                            st.session_state.categories[category],
                            key=f"remove_select_{category}"
                        )
                        if st.button(f"Remove from {category}", key=f"remove_btn_{category}"):
                            st.session_state.categories[category].remove(remove_pos)
                            assigned = sum(st.session_state.categories.values(), [])
                            st.rerun()
                    st.write("Current:", ", ".join(st.session_state.categories[category]))

            # Ensure robust repeated group generation
            if st.button("Reset Grouping Settings"):
                for key in ["categories", "auto_fill_done", "debug_data"]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()

            # Validation and group generation
            group_size = st.number_input(
                "Group Size",
                min_value=2,
                max_value=10,
                value=4,
                help="Number of participants per group",
                key="main_group_size"  # Unique key for single definition
            )
            min_members = group_size * 2

            valid = True
            for category, positions in st.session_state.categories.items():
                count = len(df[df[position_col].isin(positions)])
                if count < min_members:
                    valid = False
                    st.error(f"{category} needs at least {min_members} members (current: {count})")

            if not valid:
                if st.button("Fill Non-Filled Classes", key="auto_fill_btn"):
                    df_filled, new_categories = auto_fill_categories(
                        df,
                        position_col,
                        st.session_state.categories,
                        group_size  # Pass group_size instead of min_members
                    )

                    # Update session state with debug info
                    st.session_state.update({
                        'categories': new_categories,
                        'auto_fill_done': True,
                        'debug_data': {
                            'df_filled': df_filled.assign(
                                position_category=df_filled[position_col].apply(
                                    lambda x: next(
                                        (cat for cat, positions in new_categories.items()
                                         if x in positions),
                                        'Uncategorized'
                                    )
                                )
                            ),
                            'new_categories': new_categories
                        }
                    })

                # Permanent debug output section
                if st.session_state.get('auto_fill_done'):
                    st.success("Categories automatically filled!")

                    debug_data = st.session_state.debug_data
                    st.write("### Auto-Fill Results")

                    for category, positions in debug_data['new_categories'].items():
                        category_df = debug_data['df_filled'][debug_data['df_filled'][position_col].isin(positions)]
                        with st.expander(f"{category} ({len(category_df)} members)"):
                            # Use unique column names for display
                            display_df = category_df[[name_col, email_col, position_col]].copy()
                            display_df.columns = ['Participant Name', 'Email Address', 'Position Category']
                            st.write(display_df)

                    # In the Generate Teams button click handler:
                    if st.button("Generate Randomized Teams", key="generate_teams_btn"):
                        edited_df = df.copy()  # Ensure user-modified class assignments are used
                        for category, positions in st.session_state.categories.items():
                            edited_df.loc[edited_df[position_col].isin(positions), 'position_category'] = category
                        processor = GroupProcessor(
                            df=edited_df,
                            group_size=group_size,
                            position_col='position_category',
                            name_col=name_col,
                            email_col=email_col,
                            job_sector_col=job_sector_col
                        )
                        result = processor.generate_groups()
                        
                        if not result['success']:
                            st.error(f"Error generating groups: {result['error']}")
                        else:
                            output_df = result['df']
                            
                            # Ensure consistent column naming
                            if 'Group ID' in output_df.columns and 'Group' not in output_df.columns:
                                output_df = output_df.rename(columns={'Group ID': 'Group'})
                                
                            # Sort by group for better display
                            output_df = output_df.sort_values('Group').reset_index(drop=True)

                            import io
                            import openpyxl
                            from openpyxl.styles import PatternFill
                            from openpyxl.utils import get_column_letter
                            from pandas import ExcelWriter

                            def color_rows_by_group(writer, df):
                                wb = writer.book
                                ws = writer.sheets['Sheet1']
                                group_colors = ["FFF2CC", "D9EAD3", "CFE2F3", "F4CCCC", "D9D2E9", "C9DAF8", "EAD1DC", "B6D7A8", "FFD966"]
                                group_map = {}
                                for idx, group in enumerate(df['Group'].unique()):
                                    group_map[group] = group_colors[idx % len(group_colors)]
                                for row in range(2, len(df) + 2):
                                    group = df.iloc[row-2]['Group']
                                    fill = PatternFill(start_color=group_map[group], end_color=group_map[group], fill_type="solid")
                                    for col in range(1, len(df.columns) + 1):
                                        ws[f"{get_column_letter(col)}{row}"].fill = fill

                            excel_buffer = io.BytesIO()
                            with ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                output_df.to_excel(writer, index=False)
                                color_rows_by_group(writer, output_df)

                            st.success(f"Created {result['num_groups']} groups with {len(output_df)} participants!")
                            st.dataframe(output_df)
                            st.download_button(
                                label="Download Assignments (Excel)",
                                data=excel_buffer.getvalue(),
                                file_name="group_assignments.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            if valid and st.button("Generate Groups"):
                # Map categories to DataFrame
                edited_df = df.copy()  # Ensure user-modified class assignments are used
                for category, positions in st.session_state.categories.items():
                    edited_df.loc[edited_df[position_col].isin(positions), 'position_category'] = category

                processor = GroupProcessor(
                    df=edited_df,
                    group_size=group_size,
                    position_col='position_category',
                    name_col=name_col,
                    email_col=email_col,
                    job_sector_col=job_sector_col
                )

                result = processor.generate_groups()
                
                if not result['success']:
                    st.error(f"Error generating groups: {result['error']}")
                else:
                    output_df = result['df']
                    
                    # Ensure consistent column naming
                    if 'Group ID' in output_df.columns and 'Group' not in output_df.columns:
                        output_df = output_df.rename(columns={'Group ID': 'Group'})
                        
                    # Sort by group for better display
                    output_df = output_df.sort_values('Group').reset_index(drop=True)

                    import io
                    import openpyxl
                    from openpyxl.styles import PatternFill
                    from openpyxl.utils import get_column_letter
                    from pandas import ExcelWriter

                    def color_rows_by_group(writer, df):
                        wb = writer.book
                        ws = writer.sheets['Sheet1']
                        group_colors = ["FFF2CC", "D9EAD3", "CFE2F3", "F4CCCC", "D9D2E9", "C9DAF8", "EAD1DC", "B6D7A8", "FFD966"]
                        group_map = {}
                        for idx, group in enumerate(df['Group'].unique()):
                            group_map[group] = group_colors[idx % len(group_colors)]
                        for row in range(2, len(df) + 2):
                            group = df.iloc[row-2]['Group']
                            fill = PatternFill(start_color=group_map[group], end_color=group_map[group], fill_type="solid")
                            for col in range(1, len(df.columns) + 1):
                                ws[f"{get_column_letter(col)}{row}"].fill = fill

                    excel_buffer = io.BytesIO()
                    with ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        output_df.to_excel(writer, index=False)
                        color_rows_by_group(writer, output_df)

                    st.success(f"Created {result['num_groups']} groups with {len(output_df)} participants!")
                    st.dataframe(output_df)
                    st.download_button(
                        label="Download Assignments (Excel)",
                        data=excel_buffer.getvalue(),
                        file_name="group_assignments.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )


if __name__ == "__main__":
    main()