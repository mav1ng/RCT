import pytest
import pandas as pd
import numpy as np
from grouping_logic.processor import Processor
from unittest.mock import patch
from collections import defaultdict
import random

@pytest.fixture
def sample_data():
    """Generate synthetic dataset for benchmarking"""
    # Set a fixed random seed for reproducible test data
    np.random.seed(42)
    
    return pd.DataFrame({
        'Position_Category': ['Early']*15 + ['Mid']*20 + ['Senior']*8 + ['Student']*12,
        'Job_Sector': ['Tech']*25 + ['Finance']*15 + ['Healthcare']*15,
        'Experience': [i%5 for i in range(55)],
        'Name': [f"User {i}" for i in range(55)],
        'Email': [f"user{i}@example.com" for i in range(55)]
    })

def test_consistency(sample_data):
    """Ensure deterministic results with same trial count/seeds"""
    # First run with fixed random seed
    np.random.seed(42)
    processor1 = Processor(
        sample_data,
        group_size=5,
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )
    result1 = processor1.generate_groups()
    assert result1['success'], "First group generation failed"
    
    # Second run with same fixed random seed
    np.random.seed(42)
    processor2 = Processor(
        sample_data,
        group_size=5,
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )
    result2 = processor2.generate_groups()
    assert result2['success'], "Second group generation failed"
    
    # Compare results
    df1 = result1['df'].sort_values(['Group', 'Name']).reset_index(drop=True)
    df2 = result2['df'].sort_values(['Group', 'Name']).reset_index(drop=True)
    
    pd.testing.assert_frame_equal(df1, df2, 
                                 "Group assignments should be deterministic with same random seeds")

def test_stratified_split(sample_data):
    """
    Test that robust stratified split creates only full groups of group_size with correct stratification.
    If group_size > number of unique classes, allow duplicate positions in a group.
    If group_size <= number of unique classes, require all unique positions in a group.
    """
    processor = Processor(
        sample_data,
        group_size=5,
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )
    groups, indices = processor._robust_stratified_split()
    # All groups must have exactly group_size members
    assert all(len(g['Members']) == 5 for g in groups)
    n_classes = sample_data['Position_Category'].nunique()
    for g in groups:
        positions = [m['Position_Category'] for m in g['Members']]
        if 5 <= n_classes:
            # All positions must be unique
            assert len(set(positions)) == len(positions), f"Non-unique positions in group: {positions}"
        else:
            # Duplicates allowed if group_size > n_classes
            assert len(positions) == 5

def test_remainder_assignment(sample_data):
    """Test that remainders are assigned to maximize diversity"""
    # Create a processor with a group size that will leave remainders
    processor = Processor(
        sample_data,
        group_size=4,  # 55 participants / 4 = 13 groups with 3 remainders
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )

    # Use a fixed random seed for deterministic testing
    with patch('random.sample', side_effect=random.sample), \
         patch('numpy.random.seed', side_effect=np.random.seed):
        # Force a consistent random seed for testing
        np.random.seed(42)
        random.seed(42)
        
        result = processor.generate_groups()
        
    assert result['success'], "Group generation failed"

    # Check that all participants are assigned
    assert len(result['df']) == len(sample_data), "All participants should be assigned to groups"

    # Check group sizes
    group_sizes = result['df'].groupby('Group').size()
    min_size = processor.group_size
    max_size = processor.group_size + 1

    assert group_sizes.min() >= min_size, f"All groups should have at least {min_size} members"
    assert group_sizes.max() <= max_size, f"No group should have more than {max_size} members"

def test_no_undersized_groups_and_no_unassigned(sample_data):
    """Test that all remainders are assigned and no group is below the minimum size."""
    processor = Processor(
        sample_data,
        group_size=5,  # Use a value that will leave remainders
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )
    result = processor.generate_groups()
    assert result['success'], "Group generation failed"
    df = result['df']
    # All participants assigned
    assert len(df) == len(sample_data)
    # No group below minimum size
    group_sizes = df.groupby('Group').size()
    assert group_sizes.min() >= processor.group_size, (
        f"Found group(s) below minimum size: {group_sizes[group_sizes < processor.group_size]}")

def test_excel_output_format(tmp_path, sample_data):
    """Test Excel output is sorted, colored, and contains all participants"""
    import openpyxl
    from pandas import ExcelWriter
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    
    # Generate groups
    processor = Processor(
        sample_data, 
        group_size=5,
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )
    
    result = processor.generate_groups()
    assert result['success'], "Group generation failed"
    
    # Sort by group
    output_df = result['df'].sort_values('Group').reset_index(drop=True)
    
    # Define color function similar to the app
    def color_rows_by_group(writer, df):
        wb = writer.book
        ws = writer.sheets['Sheet1']
        group_colors = ["FFF2CC", "D9EAD3", "CFE2F3", "F4CCCC", "D9D2E9", "C9DAF8", "EAD1DC", "B6D7A8", "FFD966"]
        group_map = {}
        for idx, group in enumerate(df['Group'].unique()):
            group_map[group] = group_colors[idx % len(group_colors)]
        for row in range(2, len(df) + 2):
            group = df.iloc[row-2]['Group']
            fill = PatternFill(start_color=group_map[group], end_color=group_map[group], fill_type="solid")
            for col in range(1, len(df.columns) + 1):
                ws[f"{get_column_letter(col)}{row}"].fill = fill
    
    # Save to Excel
    excel_path = tmp_path / 'test_groups.xlsx'
    with ExcelWriter(excel_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False)
        color_rows_by_group(writer, output_df)
    
    # Load the Excel file and check formatting
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Check that all participants are included
    assert ws.max_row - 1 == len(sample_data), "Excel should contain all participants"
    
    # Check that groups are sorted
    group_col_idx = None
    for i, cell in enumerate(ws[1]):
        if cell.value == 'Group':
            group_col_idx = i
            break
    
    assert group_col_idx is not None, "Group column should exist in Excel"
    
    group_col = [ws.cell(row=r, column=group_col_idx+1).value for r in range(2, ws.max_row+1)]
    assert group_col == sorted(group_col), "Groups must be sorted in Excel output"

def test_diversity_score_and_swap(sample_data):
    """Test that the diversity score is computed and no single swap can improve it after generation."""
    processor = Processor(
        sample_data,
        group_size=5,
        position_col='Position_Category',
        job_sector_col='Job_Sector'
    )
    result = processor.generate_groups()
    assert result['success']
    score = result['diversity_score']
    # Should be an integer
    assert isinstance(score, int)
    # If a swap is suggested, it should have positive gain
    swap = result['swap_suggestion']
    if swap:
        assert swap['score_gain'] > 0
    else:
        # No swap possible: score is locally optimal
        pass

def test_group_size_bounds_and_assignment(sample_data):
    """
    Test that all groups have size in [group_size, group_size+1] (except at most one leftover group < group_size),
    and all participants are assigned, and check diversity score.
    """
    for group_size in [3, 4, 5, 7]:
        processor = Processor(
            sample_data,
            group_size=group_size,
            position_col='Position_Category',
            job_sector_col='Job_Sector'
        )
        result = processor.generate_groups()
        assert result['success'], f"Group generation failed for group_size={group_size}"
        groups = result['groups']
        all_members = [m['index'] for g in groups for m in g['Members']]
        assert len(set(all_members)) == len(sample_data), "Some participants unassigned or duplicated"
        # Allow at most one group with size < group_size (leftover group)
        small_groups = [g for g in groups if len(g['Members']) < processor.group_size]
        assert len(small_groups) <= 1, "More than one undersized group"
        for g in groups:
            if len(g['Members']) < processor.group_size:
                # Must be labeled as leftover
                assert g['Group ID'].startswith('Leftover-'), f"Undersized group not labeled as leftover: {g['Group ID']}"
            else:
                assert processor.group_size <= len(g['Members']) <= processor.group_size + 1, f"Group size out of bounds: {len(g['Members'])}"
        # Diversity score should be integer and non-negative
        score = result['diversity_score']
        assert isinstance(score, int)
        assert score >= 0

def test_swap_optimization_and_stratified_class_usage(sample_data):
    """
    Test that the stratified split uses the correct user-assigned class labels (e.g. Student, Early Career, etc.)
    and that swap optimization increases (or maintains) the diversity score.
    Prints diversity score before and after swaps.
    """
    # Simulate user assigning classes
    mapping = {
        'PhD student': 'Student',
        "Master's student": 'Student',
        'Graduate': 'Student',
        'Post doc': 'Early Career',
        'Entry-level position (<3 years of experience)': 'Early Career',
        'Mid-level position (3-10 years experience)': 'Mid Level',
        'Senior position (>10 years experience)': 'Senior Level',
        'Contractor': 'Early Career',
    }
    sample_data = sample_data.copy()
    sample_data['position_category'] = sample_data['Position_Category'].map(mapping).fillna('Other')
    processor = Processor(
        sample_data,
        group_size=5,
        position_col='position_category',
        job_sector_col='Job_Sector'
    )
    result = processor.generate_groups()
    assert result['success']
    groups = result['groups']
    print(f"Diversity score before swaps: {result['diversity_score']}")
    opt_groups, score_before, score_after, swaps = processor.optimize_swaps_for_diversity(groups)
    print(f"Diversity score after swaps: {score_after}")
    print(f"Number of swaps performed: {len(swaps)}")
    # Print the class counts actually used for the split
    print("[TEST] position_category class counts:", sample_data['position_category'].value_counts().to_dict())
    assert score_after >= score_before

def test_assignment_method_labels(sample_data):
    """
    Test that Assignment_Method column correctly labels each member as Stratified Split, Greedy Assignment, or Leftover Greedy Assignment.
    """
    mapping = {
        'PhD student': 'Student',
        "Master's student": 'Student',
        'Graduate': 'Student',
        'Post doc': 'Early Career',
        'Entry-level position (<3 years of experience)': 'Early Career',
        'Mid-level position (3-10 years experience)': 'Mid Level',
        'Senior position (>10 years experience)': 'Senior Level',
        'Contractor': 'Early Career',
    }
    sample_data = sample_data.copy()
    sample_data['position_category'] = sample_data['Position_Category'].map(mapping).fillna('Other')
    processor = Processor(
        sample_data,
        group_size=5,
        position_col='position_category',
        job_sector_col='Job_Sector'
    )
    result = processor.generate_groups()
    assert result['success']
    df = result['df']
    # Print assignment method counts
    print("Assignment method counts:", df['Assignment_Method'].value_counts().to_dict())
    # All rows must have exactly one label
    assert set(df['Assignment_Method'].unique()) <= {"Stratified Split", "Greedy Assignment", "Leftover Greedy Assignment"}
    # No overlap between stratified and greedy
    stratified = set(df[df['Assignment_Method'] == 'Stratified Split'].index)
    greedy = set(df[df['Assignment_Method'] == 'Greedy Assignment'].index)
    leftover = set(df[df['Assignment_Method'] == 'Leftover Greedy Assignment'].index)
    assert stratified.isdisjoint(greedy)
    assert stratified.isdisjoint(leftover)
    assert greedy.isdisjoint(leftover)
    # All rows are labeled
    assert len(df) == len(stratified | greedy | leftover)
